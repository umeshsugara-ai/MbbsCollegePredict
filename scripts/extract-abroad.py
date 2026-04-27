"""
Extract the three transposed abroad sheets from MBBS.xlsx and emit a single
normalised data/abroad/universities.json. The fourth tab (`Indian University`)
is intentionally skipped — it duplicates MCC CSV coverage and uses subjective
editorial scoring rather than empirical data.

Output schema per university (camelCase normalisation of the 72 source labels):
  {
    "source": "vyom_sir" | "russia_univ" | "neethu_mam",
    "name", "country", "city",
    "nmcRecognized", "whoListed", "wfmeAccredited",
    "mediumOfInstruction",
    "neetRequirement",
    "annualTuitionUSD", "hostelMessUSD", "totalProgramCostUSD",
    "annualTuitionRaw", "totalProgramCostRaw",   # original strings as authored
    "durationYears", "internshipMonths",
    "indianStudentsCount", "indianMessAvailable",
    "fmgePassRate",
    "qsRank", "country_rank",
    ... etc. — all 72 fields preserved with stable camelCase keys.
  }

Run from repo root:
  python scripts/extract-abroad.py
"""

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path


def json_default(obj):
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj).__name__} not JSON serialisable")

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "MBBS.xlsx"
OUT  = REPO / "data" / "abroad" / "universities.json"

# Sheet → source-tag mapping. Order is meaningful (output order).
ABROAD_SHEETS = [
    ("Vyom Sir",    "vyom_sir"),
    ("Russia Univ", "russia_univ"),
    ("Neethu mam",  "neethu_mam"),
]

# Stable camelCase key generation from human field labels. Strips parentheticals,
# punctuation, and unit hints — but only for the KEY. The value is preserved
# verbatim (then normalised below). Two labels that differ only in trailing
# units / parentheticals collapse to the same key, which is what we want.
def to_camel_key(label: str) -> str:
    if not label:
        return ""
    s = re.sub(r"\([^)]*\)", " ", str(label))   # strip parentheticals
    s = re.sub(r"[^A-Za-z0-9]+", " ", s).strip()
    if not s:
        return ""
    parts = s.split()
    return parts[0].lower() + "".join(w.capitalize() for w in parts[1:])

# Yes/No/Maybe normaliser. Keeps nuanced caveats untouched (e.g. Etugen's
# "Not individually approved — eligible via parent recognition") because those
# are signal, not noise.
YES_TOKENS = {"yes", "y", "✓", "true", "available", "approved", "recognized", "recognised"}
NO_TOKENS  = {"no", "n", "✗", "false", "not available", "na", "n/a", "—", "-"}

def normalise_yn(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    low = s.lower()
    # Only flatten when the entire cell IS one of these tokens. If there's any
    # extra text ("Yes — but pending review"), keep verbatim so nuance survives.
    if low in YES_TOKENS: return "Yes"
    if low in NO_TOKENS:  return "No"
    return s

# Numeric extraction for fee / cost / count fields. Returns the first number
# found. The original string is preserved separately.
def extract_number(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group()) if m else None

# USD-aware money extraction. The dataset is bilingual on currency: some cells
# say "$4500", others "USD 27,000", others "₹35 lakh", others mixed. Budget
# filtering happens server-side in USD, so we ONLY surface a `_num` value when
# we are confident the source string is USD. INR cells return None for `_num`
# (the original string is preserved on the parent key) so the server falls
# back to "let the LLM decide" rather than mis-filtering on the wrong unit.
def extract_usd(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v)   # bare numeric — assume USD context (these come from numeric cells)
    s = str(v)
    s_lower = s.lower()
    has_usd = ("$" in s) or ("usd" in s_lower)
    has_inr = ("₹" in s) or ("inr" in s_lower) or ("lakh" in s_lower) or ("lac" in s_lower)
    # If it mentions both, USD wins (the canonical figure is usually quoted
    # first in USD with INR as a parenthetical conversion). If it's INR-only,
    # don't surface a numeric — server will skip budget-filtering for this row.
    if has_inr and not has_usd:
        return None
    s_clean = s.replace(",", "")
    m = re.search(r"\$?\s*(\d+(?:\.\d+)?)", s_clean)
    return float(m.group(1)) if m else None

# Field-key heuristics: which keys should surface a numeric companion.
NUMERIC_FIELD_HINTS = (
    "fee", "cost", "tuition", "hostel", "rank", "year", "duration",
    "month", "students", "rate", "pass", "score", "index",
)
# Which keys specifically need USD-aware extraction (vs general number-only).
MONEY_FIELD_HINTS = ("fee", "cost", "tuition", "hostel", "stipend", "salary")

def is_numericish(key: str) -> bool:
    k = key.lower()
    return any(h in k for h in NUMERIC_FIELD_HINTS)

def is_moneyish(key: str) -> bool:
    k = key.lower()
    return any(h in k for h in MONEY_FIELD_HINTS)

def extract_sheet(ws, source_tag: str) -> list[dict]:
    """
    Sheet is transposed: column A = field labels, columns B..N = universities.
    Build one dict per university column. Drop columns that are 100% blank.
    """
    if ws.max_row < 2 or ws.max_column < 2:
        return []

    # Column A: field labels (rows 1..max_row). Row 1 is typically a header
    # like "University Name" or similar — keep it; it becomes a field too.
    labels = []
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1).value
        labels.append(cell)

    universities = []
    for col in range(2, ws.max_column + 1):
        # Skip the column entirely if every cell is blank.
        col_vals = [ws.cell(row=r, column=col).value for r in range(1, ws.max_row + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in col_vals):
            continue

        uni = {"source": source_tag}
        used_keys = {}  # camelKey → suffix counter for collision handling
        for label, raw in zip(labels, col_vals):
            if label is None:
                continue
            key = to_camel_key(label)
            if not key:
                continue
            # Collision: two source labels normalise to the same key. Suffix
            # the second one with _2, _3, … so we don't silently overwrite.
            if key in uni:
                used_keys[key] = used_keys.get(key, 1) + 1
                key = f"{key}_{used_keys[key]}"

            value = raw
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    value = None
                else:
                    norm = normalise_yn(value)
                    value = norm if norm is not None else value

            uni[key] = value

            # Companion numeric for fee/cost/rank-shaped keys. Suffix _num.
            # Money keys use USD-aware extraction (returns None for INR-only
            # strings) so server-side budget filtering doesn't compare USD
            # budgets against lakh-denominated numbers.
            if is_numericish(key) and value is not None and not isinstance(value, bool):
                num = extract_usd(value) if is_moneyish(key) else extract_number(value)
                if num is not None:
                    uni[f"{key}_num"] = num

        # Resolve canonical name. The 72-field schema's first row is usually
        # "University Name" / "Name of the University" → key `nameOfTheUniversity`
        # or `name`. If neither key has a non-empty value, this column is a
        # blank/template column — drop it.
        name = (
            uni.get("nameOfTheUniversity")
            or uni.get("name")
            or uni.get("universityName")
            or uni.get("nameOfUniversity")
        )
        if not name or not str(name).strip():
            continue

        # Surface a stable `name` field on every uni for server-side use.
        uni["name"] = str(name).strip()
        universities.append(uni)

    return universities


def main() -> int:
    if not XLSX.exists():
        print(f"Source file not found: {XLSX}", file=sys.stderr)
        return 1

    wb = load_workbook(XLSX, data_only=True, read_only=False)
    all_unis = []
    seen_keys: dict[str, int] = {}

    for sheet_name, source_tag in ABROAD_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"WARN: sheet '{sheet_name}' not in workbook — skipping", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        rows = extract_sheet(ws, source_tag)
        for u in rows:
            for k in u.keys():
                seen_keys[k] = seen_keys.get(k, 0) + 1
        all_unis.extend(rows)
        print(f"  {sheet_name:<14} → {len(rows):>2} universities ({source_tag})", file=sys.stderr)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "_meta": {
            "generated_from": "MBBS.xlsx",
            "sheets": [t for _, t in ABROAD_SHEETS],
            "count": len(all_unis),
            "field_coverage": dict(sorted(seen_keys.items(), key=lambda kv: -kv[1])),
        },
        "universities": all_unis,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=json_default), encoding="utf-8")
    print(f"\nWrote {len(all_unis)} universities to {OUT.relative_to(REPO)}", file=sys.stderr)
    print(f"  unique field keys: {len(seen_keys)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
